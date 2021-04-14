import os

import openpyxl
from datetime import datetime
import xlsxwriter

from openpyxl.styles import Side, Border


def update_logs(message, hmac_string, nonce, key, fail):
    dirname = os.path.dirname(__file__)
    filename = 'logs.xlsx'
    pathname = os.path.join(dirname, filename)

    if os.path.exists(pathname):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook['Logs']
        thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
        max_row = worksheet.max_row

        worksheet.cell(row=max_row+1, column=2).value = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        worksheet.cell(row=max_row+1, column=2).border = thin_border
        worksheet.cell(row=max_row+1, column=3).value = message
        worksheet.cell(row=max_row+1, column=3).border = thin_border
        worksheet.cell(row=max_row+1, column=4).value = hmac_string
        worksheet.cell(row=max_row+1, column=4).border = thin_border
        worksheet.cell(row=max_row+1, column=5).value = nonce
        worksheet.cell(row=max_row+1, column=5).border = thin_border
        worksheet.cell(row=max_row+1, column=6).value = key
        worksheet.cell(row=max_row+1, column=6).border = thin_border
        worksheet.cell(row=max_row+1, column=7).value = fail
        worksheet.cell(row=max_row+1, column=7).border = thin_border

        workbook.save(filename)

    else:
        workbook = xlsxwriter.Workbook(pathname)
        worksheet = workbook.add_worksheet('Logs')

        title_format = workbook.add_format({'bold': 1, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})
        column_title_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})
        data_format = workbook.add_format({'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})

        worksheet.set_column(1, 6, 24)

        worksheet.merge_range("B2:G2", 'Transmission integrity logs', title_format)
        worksheet.write(2, 1, 'Datetime', column_title_format)
        worksheet.write(3, 1, datetime.now().strftime("%d/%m/%Y %H:%M:%S"), data_format)

        worksheet.write(2, 2, 'Message', column_title_format)
        worksheet.write(3, 2, message, data_format)

        worksheet.write(2, 3, 'HMAC', column_title_format)
        worksheet.write(3, 3, hmac_string, data_format)

        worksheet.write(2, 4, 'Nonce', column_title_format)
        worksheet.write(3, 4, nonce, data_format)

        worksheet.write(2, 5, 'Key', column_title_format)
        worksheet.write(3, 5, key, data_format)

        worksheet.write(2, 6, 'Integrity fail', column_title_format)
        worksheet.write(3, 6, fail, data_format)

        workbook.close()
    print('SERVER INFO: Logs were updated')
